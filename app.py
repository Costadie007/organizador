import io
import math
import os
import re
import sqlite3
import urllib.parse
import zipfile
from difflib import SequenceMatcher

import cv2
import openpyxl
import pandas as pd
from openpyxl.drawing.image import Image as OpenpyxlImage
import numpy as np
from PIL import Image, ImageOps
import streamlit as st

# --- 1. BANCO DE DADOS (SQLITE) ---
DB_NAME = "sistema_usuarios.db"

def init_db():
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS usuarios (
            usuario TEXT PRIMARY KEY,
            senha TEXT NOT NULL,
            nome_completo TEXT,
            contato TEXT,
            status TEXT DEFAULT 'pendente',
            role TEXT DEFAULT 'user',
            data_cadastro TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    c.execute("SELECT * FROM usuarios WHERE usuario = ?", ("diego.costa",))
    if not c.fetchone():
        c.execute('''
            INSERT INTO usuarios (usuario, senha, nome_completo, contato, status, role)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', ("diego.costa", "admin123", "Diego Costa", "5500000000000", "aprovado", "admin"))
    conn.commit()
    conn.close()

init_db()

def buscar_usuario(usuario):
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("SELECT usuario, senha, nome_completo, contato, status, role FROM usuarios WHERE usuario = ?", (usuario.lower().strip(),))
    row = c.fetchone()
    conn.close()
    if row:
        return {
            "usuario": row[0],
            "senha": row[1],
            "nome_completo": row[2],
            "contato": row[3],
            "status": row[4],
            "role": row[5]
        }
    return None

def cadastrar_usuario(usuario, senha, nome, contato):
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    try:
        c.execute('''
            INSERT INTO usuarios (usuario, senha, nome_completo, contato, status, role)
            VALUES (?, ?, ?, ?, 'pendente', 'user')
        ''', (usuario.lower().strip(), senha, nome, contato))
        conn.commit()
        conn.close()
        return True, "Cadastro realizado com sucesso! Aguarde a aprovação do administrador."
    except sqlite3.IntegrityError:
        conn.close()
        return False, "Este nome de usuário já está cadastrado. Escolha outro."

def atualizar_status_db(usuario, novo_status):
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    if novo_status == "excluir":
        c.execute("DELETE FROM usuarios WHERE usuario = ?", (usuario,))
    else:
        c.execute("UPDATE usuarios SET status = ? WHERE usuario = ?", (novo_status, usuario))
    conn.commit()
    conn.close()

def listar_todos_usuarios():
    conn = sqlite3.connect(DB_NAME)
    df = pd.read_sql_query("SELECT usuario, nome_completo, contato, status, role, data_cadastro FROM usuarios", conn)
    conn.close()
    return df

# --- 2. MOTORES DE OCR / BARCODE ---
try:
    import zxingcpp
except ImportError:
    import zxing_cpp as zxingcpp

try:
    from pyzbar import pyzbar
    HAS_PYZBAR = True
except Exception:
    HAS_PYZBAR = False

try:
    import easyocr
    @st.cache_resource
    def carregar_ocr():
        return easyocr.Reader(['en'], gpu=False, verbose=False)
    OCR_READER = carregar_ocr()
    HAS_OCR = True
except Exception:
    HAS_OCR = False

# --- 3. CONFIGURAÇÃO VISUAL & ESTILOS PADRONIZADOS ---
st.set_page_config(
    page_title="Organizador de Planilhas",
    page_icon="📊",
    layout="wide"
)

COR_FUNDO = "#262523"
COR_CARD = "#333230"
COR_LARANJA = "#F39200"

st.markdown(f"""
    <style>
    /* Ocultar elementos padrão do Streamlit */
    [data-testid="stToolbar"], [data-testid="stHeader"], header, #MainMenu {{
        display: none !important;
        visibility: hidden !important;
    }}
    .css-1544g2n, .e16nr0p33, a.header-anchor, [data-testid="stHeaderActionElements"], .stMarkdown a[href^="#"] {{
        display: none !important;
        visibility: hidden !important;
    }}
    a[href^="#"] {{
        pointer-events: none !important;
        cursor: default !important;
        text-decoration: none !important;
    }}
    
    .stApp {{
        background-color: {COR_FUNDO};
        color: #FFFFFF;
        font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
    }}
    div.block-container {{
        padding-top: 2rem !important;
        padding-bottom: 2rem;
        max-width: 90%;
    }}
    
    /* Card de Métricas estilo Painel do Lote */
    .metric-box {{
        background-color: {COR_CARD};
        border-radius: 8px;
        padding: 16px 10px;
        text-align: center;
        margin-bottom: 12px;
        border: 1px solid rgba(255, 255, 255, 0.05);
    }}
    .metric-number {{
        font-size: 26px;
        font-weight: 800;
        color: {COR_LARANJA};
        line-height: 1;
        margin-bottom: 6px;
    }}
    .metric-label {{
        font-size: 11px;
        font-weight: 700;
        color: #A0A0A0;
        letter-spacing: 0.5px;
    }}
    
    /* Cabeçalho do Estilo Tabela para Administrador */
    .table-header {{
        background-color: {COR_CARD};
        padding: 10px 15px;
        border-radius: 6px;
        font-weight: bold;
        color: {COR_LARANJA};
        margin-bottom: 8px;
        border: 1px solid rgba(255, 255, 255, 0.05);
    }}
    .table-row {{
        padding: 8px 15px;
        border-bottom: 1px solid rgba(255, 255, 255, 0.05);
        align-items: center;
    }}

    .stButton>button {{
        background: linear-gradient(90deg, {COR_LARANJA} 0%, #d88100 100%) !important;
        color: #FFFFFF !important;
        border-radius: 6px !important;
        font-weight: bold !important;
        border: none !important;
        padding: 0.5rem 1rem !important;
    }}
    .stDownloadButton>button {{
        background-color: {COR_LARANJA} !important;
        color: #FFFFFF !important;
        border-radius: 6px !important;
        font-weight: bold !important;
        border: none !important;
    }}
    </style>
""", unsafe_allow_html=True)

# --- 4. SESSÃO E LOGIN ---
if "autenticado" not in st.session_state:
    st.session_state.autenticado = False
if "usuario_logado" not in st.session_state:
    st.session_state.usuario_logado = None
if "mensagem_aprovacao" not in st.session_state:
    st.session_state.mensagem_aprovacao = None

if not st.session_state.autenticado:
    st.markdown("<br><br>", unsafe_allow_html=True)
    c1, c2, c3 = st.columns([1, 1.6, 1])

    with c2:
        st.markdown(f"""
            <div style="text-align: center; margin-bottom: 20px;">
                <div style="color: {COR_LARANJA}; font-size: 32px; font-weight: 900;">📊 Organizador de Planilhas</div>
                <p style="color: #A0A0A0; font-size: 14px; margin-top: 5px;">Anexe suas fotos e sua planilha para preenchimento automático</p>
            </div>
        """, unsafe_allow_html=True)

        tab_login, tab_cadastro = st.tabs(["🔒 Acessar Conta", "📝 Solicitar Cadastro"])

        with tab_login:
            with st.form("form_login"):
                usr_login = st.text_input("Usuário").strip().lower()
                senha_login = st.text_input("Senha", type="password")
                btn_entrar = st.form_submit_button("Entrar", use_container_width=True)

                if btn_entrar:
                    dados_usr = buscar_usuario(usr_login)
                    if dados_usr and dados_usr["senha"] == senha_login:
                        if dados_usr["status"] == "aprovado":
                            st.session_state.autenticado = True
                            st.session_state.usuario_logado = dados_usr
                            st.rerun()
                        elif dados_usr["status"] == "pendente":
                            st.warning("⏳ Seu cadastro está aguardando aprovação do administrador.")
                        else:
                            st.error("🚫 Acesso bloqueado.")
                    else:
                        st.error("Usuário ou senha incorretos.")

        with tab_cadastro:
            with st.form("form_cadastro"):
                novo_usr = st.text_input("Usuário").strip().lower()
                nova_senha = st.text_input("Senha", type="password")
                nome_comp = st.text_input("Nome Completo")
                contato_wa = st.text_input("WhatsApp com DDD (Ex: 11999998888)")
                btn_cadastrar = st.form_submit_button("Enviar Solicitação", use_container_width=True)

                if btn_cadastrar:
                    if not novo_usr or not nova_senha or not nome_comp or not contato_wa:
                        st.error("Preencha todos os campos.")
                    else:
                        sucesso, msg = cadastrar_usuario(novo_usr, nova_senha, nome_comp, contato_wa)
                        if sucesso:
                            st.success(msg)
                        else:
                            st.error(msg)
    st.stop()

# --- BARRA LATERAL ---
usr_atual = st.session_state.usuario_logado
e_admin = (usr_atual["usuario"] == "diego.costa") or (usr_atual.get("role") == "admin")

with st.sidebar:
    st.markdown(f"👤 **{usr_atual['nome_completo']}**")
    st.caption(f"Usuário: `{usr_atual['usuario']}`")
    st.markdown("---")
    if st.button("🚪 Sair", use_container_width=True):
        st.session_state.autenticado = False
        st.session_state.usuario_logado = None
        st.rerun()

# --- 5. CABEÇALHO ---
col_logo, col_titulo = st.columns([1.2, 6])

with col_logo:
    st.markdown(f"""
        <div style="display:flex; justify-content:center; align-items:center; height:70px; background-color:{COR_CARD}; border-radius:8px;">
            <div style="color:#FFFFFF; font-size: 26px; font-weight:900; letter-spacing: 2px;">LOGO</div>
        </div>
    """, unsafe_allow_html=True)

with col_titulo:
    st.markdown("""
        <div style="font-size: 32px; font-weight: 800; color: #FFFFFF; margin: 0; line-height: 1.1;">Organizador de Planilhas</div>
        <div style="margin-top: 8px; color: #B0B0B0; font-size: 14px;">Anexe suas fotos e sua planilha abaixo e ela será preenchida automaticamente, com divisor de planilhas integrado.</div>
    """, unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)

# --- FUNÇÕES DE LEITURA E PROCESSAMENTO ---
def extrair_apenas_digitos(texto):
    if texto is None: return ""
    return str(re.sub(r'\D', '', str(texto)))

def limpar_texto_codigo(codigo_bruto):
    if not codigo_bruto: return None
    limpo = str(codigo_bruto).replace("(", "").replace(")", "").strip()
    limpo = re.sub(r'[^a-zA-Z0-9]', '', limpo)
    return limpo if len(limpo) >= 3 else None

def gerar_variacoes_imagem(cv_img):
    variacoes = []
    if cv_img is None: return variacoes
    h, w = cv_img.shape[:2]
    if max(h, w) > 1800:
        escala = 1800 / max(h, w)
        cv_img = cv2.resize(cv_img, (int(w * escala), int(h * escala)), interpolation=cv2.INTER_AREA)

    variacoes.append(cv_img)
    cinza = cv2.cvtColor(cv_img, cv2.COLOR_BGR2GRAY) if len(cv_img.shape) == 3 else cv_img
    variacoes.append(cinza)

    clahe = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(8, 8))
    variacoes.append(clahe.apply(cinza))

    nitida = cv2.filter2D(cinza, -1, np.array([[0, -1, 0], [-1, 5, -1], [0, -1, 0]]))
    variacoes.append(nitida)

    _, otsu = cv2.threshold(cinza, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    variacoes.append(otsu)
    return variacoes

def tentar_decodificar_leitores(img_np):
    codigos = set()
    if img_np is None: return codigos
    try:
        for res in zxingcpp.read_barcodes(img_np):
            c = limpar_texto_codigo(res.text)
            if c: codigos.add(c)
    except Exception: pass

    if HAS_PYZBAR:
        try:
            for obj in pyzbar.decode(img_np):
                c = limpar_texto_codigo(obj.data.decode("utf-8", errors="ignore"))
                if c: codigos.add(c)
        except Exception: pass
    return codigos

def tentar_ocr_extremo(img_np):
    codigos = set()
    if not HAS_OCR or img_np is None: return codigos
    try:
        res = OCR_READER.readtext(img_np, allowlist='0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZ-')
        for bbox, texto, confianca in res:
            if confianca > 0.15:
                texto_corr = (texto.replace('O', '0').replace('I', '1').replace('L', '1')
                              .replace('Z', '2').replace('S', '5').replace('B', '8').replace('G', '6'))
                c = limpar_texto_codigo(texto_corr)
                if c and len(c) >= 3: codigos.add(c)
    except Exception: pass
    return codigos

def ler_imagem_todas_camadas(cv_img, nome_arquivo):
    codigos_encontrados = set()
    digs_nome = extrair_apenas_digitos(nome_arquivo)
    if digs_nome and len(digs_nome) >= 4:
        codigos_encontrados.add(digs_nome)

    if cv_img is None: return list(codigos_encontrados)

    orientacoes = [
        cv_img,
        cv2.rotate(cv_img, cv2.ROTATE_90_CLOCKWISE),
        cv2.rotate(cv_img, cv2.ROTATE_180),
        cv2.rotate(cv_img, cv2.ROTATE_90_COUNTERCLOCKWISE)
    ]

    for img_rot in orientacoes:
        variacoes = gerar_variacoes_imagem(img_rot)
        for var in variacoes:
            cods = tentar_decodificar_leitores(var)
            if cods: codigos_encontrados.update(cods)

        if codigos_encontrados: break

        if HAS_OCR:
            for var in variacoes[:3]:
                cods_ocr = tentar_ocr_extremo(var)
                if cods_ocr: codigos_encontrados.update(cods_ocr)
            if codigos_encontrados: break

    return list(codigos_encontrados)

def calcular_similaridade_avancada(digitos_excel, texto_foto_completo, digitos_foto_lista):
    if not digitos_excel: return 0.0
    digs_foto_concat = "".join(digitos_foto_lista) + extrair_apenas_digitos(texto_foto_completo)

    if digitos_excel in digs_foto_concat: return 1.0

    tam_ex = len(digitos_excel)
    if tam_ex >= 4 and len(digs_foto_concat) >= tam_ex:
        melhor_ratio = 0.0
        for i in range(len(digs_foto_concat) - tam_ex + 1):
            sub = digs_foto_concat[i:i+tam_ex]
            ratio = SequenceMatcher(None, digitos_excel, sub).ratio()
            if ratio > melhor_ratio: melhor_ratio = ratio
        if melhor_ratio >= 0.75: return melhor_ratio

    if tam_ex >= 4:
        sufixo_4 = digitos_excel[-4:]
        sufixo_6 = digitos_excel[-6:] if tam_ex >= 6 else sufixo_4
        if sufixo_6 in digs_foto_concat: return 0.90
        if sufixo_4 in digs_foto_concat: return 0.75

    return SequenceMatcher(None, digitos_excel, digs_foto_concat).ratio()

# --- 6. ABAS PRINCIPAIS ---
if e_admin:
    tab_ferramenta, tab_admin = st.tabs(["📊 Ferramenta de Organização", "👑 Painel do Administrador"])
else:
    tab_ferramenta = st.tabs(["📊 Ferramenta de Organização"])[0]
    tab_admin = None

# ==========================================
# ABA 1: FERRAMENTA DE ORGANIZAÇÃO
# ==========================================
with tab_ferramenta:
    col_left, col_right = st.columns([2.3, 1])

    with col_left:
        st.markdown("<div style='font-size: 15px; font-weight: 600; color: #FFFFFF; margin-bottom: 8px;'>1. Selecione a Planilha Excel (.xlsx)</div>", unsafe_allow_html=True)
        arquivo_excel = st.file_uploader("Upload Planilha", type=["xlsx"], label_visibility="collapsed")

        st.markdown("<div style='font-size: 15px; font-weight: 600; color: #FFFFFF; margin-top: 15px; margin-bottom: 8px;'>2. Selecione ou arraste o lote de fotos (ou .ZIP)</div>", unsafe_allow_html=True)
        arquivos_fotos = st.file_uploader("Upload Fotos", type=["jpg", "jpeg", "png", "zip"], accept_multiple_files=True, label_visibility="collapsed")

    planilha_carregada = "Sim" if arquivo_excel else "Não"
    qtd_fotos = len(arquivos_fotos) if arquivos_fotos else 0
    vinculos_encontrados = len(st.session_state.mapa_codigo_imagem) if "mapa_codigo_imagem" in st.session_state else 0

    with col_right:
        st.markdown("<div style='font-size: 16px; font-weight: 700; color: #FFFFFF; margin-bottom: 12px;'>📊 Painel do Lote</div>", unsafe_allow_html=True)
        
        st.markdown(f"""
            <div class="metric-box">
                <div class="metric-number">{planilha_carregada}</div>
                <div class="metric-label">PLANILHA CARREGADA</div>
            </div>
        """, unsafe_allow_html=True)

        st.markdown(f"""
            <div class="metric-box">
                <div class="metric-number">{qtd_fotos}</div>
                <div class="metric-label">FOTOS CARREGADAS</div>
            </div>
        """, unsafe_allow_html=True)

        st.markdown(f"""
            <div class="metric-box">
                <div class="metric-number">{vinculos_encontrados}</div>
                <div class="metric-label">VÍNCULOS ENCONTRADOS</div>
            </div>
        """, unsafe_allow_html=True)

    if arquivo_excel and arquivos_fotos:
        st.markdown("---")
        st.markdown("<div style='font-size: 18px; font-weight: 700; color: #FFFFFF; margin-bottom: 12px;'>3. Mapeamento e Parâmetros</div>", unsafe_allow_html=True)

        wb_temp = openpyxl.load_workbook(arquivo_excel)
        nome_aba = st.selectbox("Selecione a aba da planilha:", wb_temp.sheetnames)
        df_temp = pd.read_excel(arquivo_excel, sheet_name=nome_aba)

        col1, col2 = st.columns(2)
        with col1:
            coluna_sgp = st.selectbox("Coluna com o código SGP/Principal:", list(df_temp.columns))
        with col2:
            nome_coluna_foto = st.text_input("Nome da coluna para INSERIR A FOTO:", value="FOTO")

        if st.button("🚀 INICIAR PROCESSAMENTO AUTOMÁTICO", use_container_width=True):
            st.session_state.mapa_codigo_imagem = {}

            lista_fotos = []
            for f in arquivos_fotos:
                if f.name.endswith(".zip"):
                    with zipfile.ZipFile(f) as z:
                        for filename in z.namelist():
                            if filename.lower().endswith(('.png', '.jpg', '.jpeg')):
                                img_bytes = z.read(filename)
                                pil_img = ImageOps.exif_transpose(Image.open(io.BytesIO(img_bytes)))
                                lista_fotos.append((filename, pil_img, img_bytes))
                else:
                    img_bytes = f.read()
                    pil_img = ImageOps.exif_transpose(Image.open(io.BytesIO(img_bytes)))
                    lista_fotos.append((f.name, pil_img, img_bytes))

            banco_fotos = []
            container_status = st.container()

            with container_status:
                status_leitura = st.empty()
                prog_bar = st.progress(0)
                status_leitura.write(f"🔍 Analisando {len(lista_fotos)} imagens com IA e Leitura de Códigos...")

                for idx, (nome_f, pil_img, img_bytes) in enumerate(lista_fotos):
                    np_arr = np.frombuffer(img_bytes, np.uint8)
                    cv_img = cv2.imdecode(np_arr, cv2.IMREAD_COLOR)
                    codigos_extraidos = ler_imagem_todas_camadas(cv_img, nome_f)

                    banco_fotos.append({
                        "nome": nome_f,
                        "pil_img": pil_img,
                        "codigos": codigos_extraidos,
                        "digitos_list": [extrair_apenas_digitos(c) for c in codigos_extraidos if extrair_apenas_digitos(c)],
                        "usada": False
                    })
                    prog_bar.progress((idx + 1) / len(lista_fotos))
                prog_bar.empty()

            vincularam_auto = 0
            codigos_excel_dict = {extrair_apenas_digitos(val): val for val in df_temp[coluna_sgp].dropna().unique() if extrair_apenas_digitos(val)}

            for cod_excel, val_raw in codigos_excel_dict.items():
                for foto in banco_fotos:
                    if foto["usada"]: continue
                    if calcular_similaridade_avancada(cod_excel, foto["nome"], foto["digitos_list"]) >= 0.85:
                        st.session_state.mapa_codigo_imagem[cod_excel] = foto["pil_img"]
                        foto["usada"] = True
                        vincularam_auto += 1
                        break

            vincularam_fuzzy = 0
            for cod_excel, val_raw in codigos_excel_dict.items():
                if cod_excel in st.session_state.mapa_codigo_imagem: continue
                melhor_score, melhor_foto = 0.0, None
                for foto in banco_fotos:
                    if foto["usada"]: continue
                    score = calcular_similaridade_avancada(cod_excel, foto["nome"], foto["digitos_list"])
                    if score > melhor_score and score >= 0.45:
                        melhor_score, melhor_foto = score, foto

                if melhor_foto is not None:
                    st.session_state.mapa_codigo_imagem[cod_excel] = melhor_foto["pil_img"]
                    melhor_foto["usada"] = True
                    vincularam_fuzzy += 1

            st.success(f"🎉 Processamento concluído! {len(st.session_state.mapa_codigo_imagem)} fotos vinculadas com sucesso.")
            st.rerun()

        if "mapa_codigo_imagem" in st.session_state and len(st.session_state.mapa_codigo_imagem) > 0:
            st.markdown("---")
            st.markdown("<div style='font-size: 18px; font-weight: 700; color: #FFFFFF; margin-bottom: 12px;'>4. Configuração de Saída e Divisão</div>", unsafe_allow_html=True)
            modo_divisao = st.radio("Como deseja salvar a planilha?", ["Planilha Única (Sem divisão)", "Dividir por QUANTIDADE DE PARTES", "Dividir por QUANTIDADE DE LINHAS"], horizontal=True)

            num_partes, linhas_por_parte = 1, 0
            if modo_divisao == "Dividir por QUANTIDADE DE PARTES":
                num_partes = st.number_input("Quantidade de partes:", min_value=2, value=2, step=1)
            elif modo_divisao == "Dividir por QUANTIDADE DE LINHAS":
                linhas_por_parte = st.number_input("Número de linhas por arquivo:", min_value=10, value=150, step=10)

            if st.button("📊 APLICAR FOTOS E GERAR ARQUIVO(S)", use_container_width=True):
                arquivo_excel.seek(0)
                wb = openpyxl.load_workbook(arquivo_excel)
                ws = wb[nome_aba]

                col_idx_codigo = next(col for col in range(1, ws.max_column + 1) if str(ws.cell(row=1, column=col).value or "").strip() == str(coluna_sgp).strip())
                col_idx_foto = ws.max_column + 1
                ws.cell(row=1, column=col_idx_foto).value = nome_coluna_foto

                # Configurações de dimensão padronizada para as fotos no Excel
                LARGURA_FOTO_PX = 160
                ALTURA_FOTO_PX = 80
                ALTURA_LINHA_EXCEL = 65
                LARGURA_COLUNA_EXCEL = 24

                col_letter = openpyxl.utils.get_column_letter(col_idx_foto)
                ws.column_dimensions[col_letter].width = LARGURA_COLUNA_EXCEL

                mapa_fotos_linha, vincularam = {}, 0
                for row in range(2, ws.max_row + 1):
                    cod_num = extrair_apenas_digitos(ws.cell(row=row, column=col_idx_codigo).value)
                    if cod_num and cod_num in st.session_state.mapa_codigo_imagem:
                        pil_img = st.session_state.mapa_codigo_imagem[cod_num].copy()
                        
                        # Padroniza dimensão da imagem mantendo a proporção de forma nítida
                        pil_img.thumbnail((LARGURA_FOTO_PX, ALTURA_FOTO_PX), Image.LANCZOS)
                        
                        img_byte_arr = io.BytesIO()
                        pil_img.save(img_byte_arr, format='PNG')
                        img_byte_arr.seek(0)

                        img_excel = OpenpyxlImage(img_byte_arr)
                        img_excel.width = pil_img.width
                        img_excel.height = pil_img.height

                        ws.add_image(img_excel, f"{col_letter}{row}")
                        ws.row_dimensions[row].height = ALTURA_LINHA_EXCEL
                        
                        vincularam += 1
                        mapa_fotos_linha[row] = img_byte_arr

                total_dados = ws.max_row - 1

                if modo_divisao == "Planilha Única (Sem divisão)" or total_dados <= 0:
                    out_buffer = io.BytesIO()
                    wb.save(out_buffer)
                    out_buffer.seek(0)
                    st.download_button("📥 BAIXAR PLANILHA COMPLETA (.XLSX)", out_buffer, "Planilha_Com_Fotos.xlsx", use_container_width=True)
                else:
                    if modo_divisao == "Dividir por QUANTIDADE DE PARTES":
                        linhas_por_parte = math.ceil(total_dados / num_partes)
                    qtd_partes_calculadas = math.ceil(total_dados / linhas_por_parte)

                    zip_buffer = io.BytesIO()
                    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
                        for p in range(qtd_partes_calculadas):
                            dado_inicio = (p * linhas_por_parte) + 1
                            dado_fim = min(total_dados, (p + 1) * linhas_por_parte)
                            linha_orig_inicio, linha_orig_fim = dado_inicio + 1, dado_fim + 1

                            arquivo_excel.seek(0)
                            wb_p = openpyxl.load_workbook(arquivo_excel)
                            ws_p = wb_p[nome_aba]
                            ws_p._images.clear()

                            if ws_p.max_row > linha_orig_fim: ws_p.delete_rows(linha_orig_fim + 1, ws_p.max_row - linha_orig_fim)
                            if linha_orig_inicio > 2: ws_p.delete_rows(2, linha_orig_inicio - 2)

                            col_f_p = ws_p.max_column + 1
                            ws_p.cell(row=1, column=col_f_p).value = nome_coluna_foto
                            col_letter_p = openpyxl.utils.get_column_letter(col_f_p)
                            ws_p.column_dimensions[col_letter_p].width = LARGURA_COLUNA_EXCEL

                            l_dest = 2
                            for l_orig in range(linha_orig_inicio, linha_orig_fim + 1):
                                if l_orig in mapa_fotos_linha:
                                    img_b = mapa_fotos_linha[l_orig]
                                    img_b.seek(0)
                                    
                                    # Ajuste padronizado para as partes divididas
                                    pil_p = Image.open(img_b)
                                    img_excel_p = OpenpyxlImage(img_b)
                                    img_excel_p.width = pil_p.width
                                    img_excel_p.height = pil_p.height

                                    ws_p.add_image(img_excel_p, f"{col_letter_p}{l_dest}")
                                    ws_p.row_dimensions[l_dest].height = ALTURA_LINHA_EXCEL
                                l_dest += 1

                            out_p = io.BytesIO()
                            wb_p.save(out_p)
                            zf.writestr(f"Planilha_Parte_{p+1}_(Linhas_{dado_inicio}-{dado_fim}).xlsx", out_p.getvalue())

                    zip_buffer.seek(0)
                    st.download_button("📥 BAIXAR PACOTE DE PLANILHAS (.ZIP)", zip_buffer, "Planilhas_Divididas.zip", use_container_width=True)

# ==========================================
# ABA 2: PAINEL DO ADMINISTRADOR (LAYOUT TABELA)
# ==========================================
if e_admin and tab_admin:
    with tab_admin:
        st.markdown("<div style='font-size: 20px; font-weight: bold; color: #FFFFFF; margin-bottom: 12px;'>👑 Gestão de Usuários e Aprovações</div>", unsafe_allow_html=True)
        
        # Alerta com botão para WhatsApp
        if st.session_state.mensagem_aprovacao:
            st.success(st.session_state.mensagem_aprovacao["texto"])
            st.markdown(f"""
                <a href="{st.session_state.mensagem_aprovacao['link_wa']}" target="_blank" style="
                    display: inline-block;
                    background-color: #25D366;
                    color: white;
                    padding: 8px 16px;
                    text-decoration: none;
                    border-radius: 6px;
                    font-weight: bold;
                    margin-bottom: 15px;">
                    💬 Enviar Notificação via WhatsApp
                </a>
            """, unsafe_allow_html=True)
            if st.button("Fechar Notificação"):
                st.session_state.mensagem_aprovacao = None
                st.rerun()

        df_usuarios = listar_todos_usuarios()

        # Cabeçalho no Estilo Tabela
        c_nome, c_usr, c_contato, c_status, c_role, c_acao = st.columns([2.5, 1.8, 1.8, 1.2, 1, 2.2])
        c_nome.markdown("<div class='table-header'>Nome Completo</div>", unsafe_allow_html=True)
        c_usr.markdown("<div class='table-header'>Usuário</div>", unsafe_allow_html=True)
        c_contato.markdown("<div class='table-header'>Contato</div>", unsafe_allow_html=True)
        c_status.markdown("<div class='table-header'>Status</div>", unsafe_allow_html=True)
        c_role.markdown("<div class='table-header'>Função</div>", unsafe_allow_html=True)
        c_acao.markdown("<div class='table-header'>Ação</div>", unsafe_allow_html=True)

        # Linhas da Tabela
        for idx, user in df_usuarios.iterrows():
            col1, col2, col3, col4, col5, col6 = st.columns([2.5, 1.8, 1.8, 1.2, 1, 2.2])
            
            col1.write(f"**{user['nome_completo']}**")
            col2.write(f"`{user['usuario']}`")
            col3.write(f"{user['contato']}")
            
            status_tag = "🟢 Aprovado" if user['status'] == 'aprovado' else "⏳ Pendente"
            col4.write(status_tag)
            col5.write(f"`{user['role']}`")

            # Coluna de Ações em linha
            with col6:
                if user['status'] == 'pendente':
                    btn_col1, btn_col2 = st.columns(2)
                    
                    num_limpo = re.sub(r'\D', '', str(user['contato']))
                    msg = urllib.parse.quote(f"Olá {user['nome_completo']}! Seu cadastro no Organizador de Planilhas foi APROVADO com sucesso. Você já pode fazer login e acessar o sistema!")
                    link_wa = f"https://wa.me/{num_limpo}?text={msg}"

                    if btn_col1.button("✅", key=f"ap_{user['usuario']}", help="Aprovar Usuário"):
                        atualizar_status_db(user['usuario'], 'aprovado')
                        st.session_state.mensagem_aprovacao = {
                            "texto": f"Usuário **{user['nome_completo']}** aprovado!",
                            "link_wa": link_wa
                        }
                        st.rerun()
                        
                    if btn_col2.button("🚫", key=f"rec_{user['usuario']}", help="Recusar Solicitação"):
                        atualizar_status_db(user['usuario'], 'excluir')
                        st.rerun()
                else:
                    # Para usuários já aprovados/cadastrados
                    if user['usuario'] != "diego.costa":
                        if st.button("🗑️ Excluir", key=f"del_{user['usuario']}", help="Excluir Usuário do Banco de Dados"):
                            atualizar_status_db(user['usuario'], 'excluir')
                            st.success(f"Usuário @{user['usuario']} excluído!")
                            st.rerun()
                    else:
                        st.caption("👑 Admin Principal")

            st.markdown("<div style='margin-bottom: 5px;'></div>", unsafe_allow_html=True)

# --- 7. RODAPÉ ---
st.markdown("<br><br><hr style='border-color: rgba(255, 255, 255, 0.1);'>", unsafe_allow_html=True)
st.markdown("<div style='text-align:center; color:#777777; font-size: 13px;'>Desenvolvido por <strong>Diego Costa</strong></div>", unsafe_allow_html=True)
